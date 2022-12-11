from flask import Flask, request, jsonify, redirect, url_for, render_template, session
import pickle
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import os
import git

app = Flask(__name__)

app.secret_key = "hello"


@app.route('/git_update', methods=['POST'])
def git_update():
    repo = git.Repo('./PCOS-Diagnosing-SVM-DT')
    origin = repo.remotes.origin
    repo.create_head('main',
                     origin.refs.main).set_tracking_branch(origin.refs.main).checkout()
    origin.pull()
    return '', 200

@app.route("/")
def home():
    return render_template("index.html")


THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
app.config["EXCEL_UPLOADS"] = "static/assets/uploads"
my_excel = os.path.join(THIS_FOLDER, "static/assets/uploads")
app.config["ASSETS"] = "static/assets"
my_assets = os.path.join(THIS_FOLDER, "static/assets")
app.config["ALLOWED_EXCEL_EXTENSIONS"] = ["XLSX", "CSV", "XLS"]

def predict_excel_svm(excel):
    wb = load_workbook(excel)

    ws = wb.active

    PatID = ws["A2"].value
    Age = ws["B2"].value
    Hairgrowth = ws["I2"].value
    SkinDarkening = ws["J2"].value
    PulseRateBPM = ws["Q2"].value
    CycleRI = ws["T2"].value
    FSHmIUmL = ws["AA2"].value
    LHmIUmL = ws["AB2"].value
    AMHngmL = ws["AE2"].value
    PRGngmL = ws["AH2"].value
    RBSmgdl = ws["AI2"].value
    BP_SystolicmmHg = ws["AJ2"].value
    BP_DiastolicmmHg = ws["AK2"].value
    AvgFsizeLmm = ws["AN2"].value
    AvgFsizeRmm = ws["AO2"].value
    Endometriummm = ws["AP2"].value

    session["PatID"] = PatID
    session["Age"] = Age
    session["Hairgrowth"] = Hairgrowth
    session["CycleRI"] = CycleRI
    session["AvgFsizeLmm"] = AvgFsizeLmm
    session["AvgFsizeRmm"] = AvgFsizeRmm


    model = pickle.load(open(os.path.join(my_assets, "svm-model.pkl"), 'rb'))
    session['model'] = "SVM"


    makeprediction = model.predict([[Age, Hairgrowth, SkinDarkening,
                                    PulseRateBPM, CycleRI, FSHmIUmL, LHmIUmL,
                                    AMHngmL, PRGngmL, RBSmgdl, BP_SystolicmmHg,
                                    BP_DiastolicmmHg, AvgFsizeLmm, AvgFsizeRmm, Endometriummm]])

    output = round(makeprediction[0], 2)

    return(output)

def predict_excel_dt(excel):
    wb = load_workbook(excel)

    ws = wb.active

    PatID = ws["A2"].value
    Age = ws["B2"].value
    Hairgrowth = ws["I2"].value
    SkinDarkening = ws["J2"].value
    PulseRateBPM = ws["Q2"].value
    CycleRI = ws["T2"].value
    FSHmIUmL = ws["AA2"].value
    LHmIUmL = ws["AB2"].value
    AMHngmL = ws["AE2"].value
    PRGngmL = ws["AH2"].value
    RBSmgdl = ws["AI2"].value
    BP_SystolicmmHg = ws["AJ2"].value
    BP_DiastolicmmHg = ws["AK2"].value
    AvgFsizeLmm = ws["AN2"].value
    AvgFsizeRmm = ws["AO2"].value
    Endometriummm = ws["AP2"].value

    session["PatID"] = PatID
    session["Age"] = Age
    session["Hairgrowth"] = Hairgrowth
    session["CycleRI"] = CycleRI
    session["AvgFsizeLmm"] = AvgFsizeLmm
    session["AvgFsizeRmm"] = AvgFsizeRmm


    model = pickle.load(open(os.path.join(my_assets, "dt-model.pkl"), 'rb'))
    session['model'] = "DT"


    makeprediction = model.predict([[Age, Hairgrowth, SkinDarkening,
                                    PulseRateBPM, CycleRI, FSHmIUmL, LHmIUmL,
                                    AMHngmL, PRGngmL, RBSmgdl, BP_SystolicmmHg,
                                    BP_DiastolicmmHg, AvgFsizeLmm, AvgFsizeRmm, Endometriummm]])

    output = round(makeprediction[0], 2)

    return(output)

def allowed_excel(filename):

    if not "." in filename:
        return False

    ext = filename.rsplit(".", 1)[1]

    if ext.upper() in app.config["ALLOWED_EXCEL_EXTENSIONS"]:
        return True
    else:
        return False

@app.route("/tool", methods=["GET", "POST"])
def tool():
    session.pop("result", None)
    session.pop("model", None)
    if request.method == "POST":
        if request.files:
            excel = request.files["input"]

            if excel.filename == "":
                print("Excel file must have a filename")
                return redirect(request.url)

            if not allowed_excel(excel.filename):
                print("That excel extension is not allowed")
                return redirect(request.url)

            else:
                filename = secure_filename(excel.filename)
                excel.save(os.path.join(my_excel, filename))
                session['save_excel'] = filename

            output = predict_excel_svm(excel)
            session['result'] = int(output)
            
            return redirect(url_for("result"))
    else:    
        if "result" in session:
            return redirect(url_for("pop"))
    return render_template("tool.html")


@app.route("/dt", methods=["GET", "POST"])
def dt():
    session.pop("result", None)
    session.pop("model", None)
    if request.method == "POST":
        if request.files:
            excel = request.files["input"]

            if excel.filename == "":
                print("Excel file must have a filename")
                return redirect(request.url)

            if not allowed_excel(excel.filename):
                print("That excel extension is not allowed")
                return redirect(request.url)

            else:
                filename = secure_filename(excel.filename)
                excel.save(os.path.join(my_excel, filename))
                session['save_excel'] = filename

            output = predict_excel_dt(excel)
            session['result'] = int(output)
            
            
            return redirect(url_for("result"))
    else:    
        if "result" in session:
            return redirect(url_for("pop"))
    return render_template("dt.html")


@app.route("/result", methods=["GET", "POST"])
def result():

    save_excel = session['save_excel']
    book = load_workbook(open(os.path.join(my_excel, save_excel), 'rb'))
    sheet = book.active
    
    if "result" in session:
        result = session["result"]
        model = session["model"]
        PatID = session["PatID"]
        Age = session["Age"]
        Hairgrowth = session["Hairgrowth"]
        CycleRI = session["CycleRI"]
        AvgFsizeLmm = session["AvgFsizeLmm"]
        AvgFsizeRmm = session["AvgFsizeRmm"]
        if model == "SVM":
            model_name = "SVM"
        else:
            model_name = "DT"
            
        print(result)
        if result == 1:
            return render_template("results.html", RESULTS="POSITIVE", EXCEL=sheet, MODEL=model_name, ID=PatID, AGE=Age, HAIR=Hairgrowth, CYC=CycleRI, AFL=AvgFsizeLmm, AFR=AvgFsizeRmm)
        else:
            return render_template("results.html", RESULTS="NEGATIVE", EXCEL=sheet, MODEL=model_name, ID=PatID, AGE=Age, HAIR=Hairgrowth, CYC=CycleRI, AFL=AvgFsizeLmm, AFR=AvgFsizeRmm)
    else:
        return redirect(url_for("tool"))

@app.route("/about")
def about():
    return render_template("about.html")

@app.route("/pop")
def pop():
    session.pop("result", None)
    session.pop("model", None)
    return redirect(url_for("tool"))

if __name__ == "__main__":
    app.run(debug=True)
